from __future__ import annotations

import argparse
import ast
import importlib
import json
import platform
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


DEFAULT_WORKBOOK = Path(r"C:\Users\AnasDahou\Downloads\srm_ep_only.xlsx")
DEFAULT_MODELS = Path(
    r"C:\Users\AnasDahou\Desktop\srm_collecte\API_GeoDjango\pprcollecte\api\models.py"
)
DEFAULT_REPORT = Path(
    r"C:\Users\AnasDahou\Desktop\srm_collecte\docs\srm_ep_dictionary_audit.md"
)
DEFAULT_JSON = Path(
    r"C:\Users\AnasDahou\Desktop\srm_collecte\docs\srm_ep_dictionary_audit.json"
)


WORKBOOK_TO_MODEL_TABLE_ALIASES = {
    "ep_bf": "borne_fontaine",
    "ep_bouche_arro": "bouche_darrosage",
    "ep_branchement": "branchement",
    "ep_brc_pt": "compteur_abonne",
    "ep_compteur_i": "compteur_reseau",
    "ep_conduite": "ep_conduite_bureau",
    "ep_cone_reduc": "cone_de_reduction",
    "ep_forage": "forage",
    "ep_hydrant": "hydrant",
    "ep_noeud": "noeud",
    "ep_obturateur": "obturateur",
    "ep_pompe": "pompe",
    "ep_puit": "puit",
    "ep_regard": "regard_ep",
    "ep_reduc_pres": "reducteur_de_pression",
    "ep_reservoir": "reservoir",
    "ep_station_pompage": "station_de_pompage",
    "ep_traversee": "traverse",
    "ep_vanne": "vanne",
    "ep_ventouse": "ventouse",
    "ep_vidange": "vanne_de_vidange",
    "conduite_terrain": "ep_conduite_terrain",
}


@dataclass
class WorkbookTable:
    schema: str
    table: str
    geometry: str | None
    dimension: str | None
    srid: str | None
    sheet_name: str | None
    fields: list[str]


@dataclass
class ModelTable:
    class_name: str
    schema: str
    table: str
    fields: list[str]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compare the SRM workbook dictionary against Django models."
    )
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--models", type=Path, default=DEFAULT_MODELS)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--json", type=Path, default=DEFAULT_JSON)
    parser.add_argument("--schema", default="ep")
    return parser.parse_args()


def safe_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_label(value: Any) -> str:
    text = safe_str(value) or ""
    normalized = unicodedata.normalize("NFKD", text)
    ascii_only = normalized.encode("ascii", "ignore").decode("ascii")
    return ascii_only.strip().lower()


def parse_workbook_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, Any]:
    schema = None
    table = None
    header_seen = False
    fields: list[str] = []

    for row in ws.iter_rows(values_only=True):
        first = safe_str(row[0] if row else None)
        third = safe_str(row[2] if len(row) > 2 else None)
        label = normalize_label(first)
        if label == "schema":
            schema = third
        elif label == "table":
            table = third
        elif label == "attribut":
            header_seen = True
            continue
        elif header_seen and first:
            fields.append(first)

    return {
        "sheet_name": ws.title,
        "schema": schema,
        "table": table,
        "fields": fields,
    }


def parse_workbook_dictionary(workbook_path: Path, target_schema: str) -> dict[str, WorkbookTable]:
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_details = {}
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        details = parse_workbook_sheet(ws)
        if details["schema"] == target_schema and details["table"]:
            sheet_details[details["table"]] = details

    tables_sheet = workbook["tables"]
    output: dict[str, WorkbookTable] = {}
    for row in tables_sheet.iter_rows(min_row=3, values_only=True):
        schema = safe_str(row[0] if len(row) > 0 else None)
        table = safe_str(row[1] if len(row) > 1 else None)
        if not schema or not table or schema != target_schema:
            continue
        geometry = safe_str(row[2] if len(row) > 2 else None)
        dimension = safe_str(row[3] if len(row) > 3 else None)
        srid = safe_str(row[4] if len(row) > 4 else None)
        details = sheet_details.get(table, {})
        output[table] = WorkbookTable(
            schema=schema,
            table=table,
            geometry=geometry,
            dimension=dimension,
            srid=srid,
            sheet_name=details.get("sheet_name"),
            fields=details.get("fields", []),
        )
    return output


def parse_db_table(raw: str) -> tuple[str | None, str]:
    cleaned = raw.replace('"', "").replace("'", "").strip()
    if "." in cleaned:
        schema, table = cleaned.split(".", 1)
        return schema, table
    return None, cleaned


def extract_model_tables(models_path: Path, target_schema: str) -> dict[str, ModelTable]:
    source = models_path.read_text(encoding="utf-8", errors="replace")
    module = ast.parse(source)
    output: dict[str, ModelTable] = {}

    for node in module.body:
        if not isinstance(node, ast.ClassDef):
            continue

        fields: list[str] = []
        db_table = None
        for item in node.body:
            if (
                isinstance(item, ast.Assign)
                and len(item.targets) == 1
                and isinstance(item.targets[0], ast.Name)
                and isinstance(item.value, ast.Call)
            ):
                func = item.value.func
                if isinstance(func, ast.Attribute) and func.attr.endswith("Field"):
                    fields.append(item.targets[0].id)
            elif isinstance(item, ast.ClassDef) and item.name == "Meta":
                for meta_item in item.body:
                    if (
                        isinstance(meta_item, ast.Assign)
                        and len(meta_item.targets) == 1
                        and isinstance(meta_item.targets[0], ast.Name)
                        and meta_item.targets[0].id == "db_table"
                        and isinstance(meta_item.value, ast.Constant)
                        and isinstance(meta_item.value.value, str)
                    ):
                        db_table = meta_item.value.value

        if not db_table:
            continue
        schema, table = parse_db_table(db_table)
        if schema != target_schema:
            continue
        output[table] = ModelTable(
            class_name=node.name,
            schema=schema or "",
            table=table,
            fields=fields,
        )
    return output


def build_comparison(
    workbook_tables: dict[str, WorkbookTable],
    model_tables: dict[str, ModelTable],
) -> dict[str, Any]:
    exact_matches = sorted(set(workbook_tables) & set(model_tables))
    mapped_matches = []
    unresolved_workbook = []

    for workbook_table in sorted(workbook_tables):
        if workbook_table in model_tables:
            continue
        mapped_table = WORKBOOK_TO_MODEL_TABLE_ALIASES.get(workbook_table)
        if mapped_table and mapped_table in model_tables:
            mapped_matches.append((workbook_table, mapped_table))
        else:
            unresolved_workbook.append(workbook_table)

    resolved_model_tables = set(exact_matches) | {model for _, model in mapped_matches}
    unresolved_models = sorted(set(model_tables) - resolved_model_tables)

    table_diffs = []
    for workbook_table, model_table in (
        [(name, name) for name in exact_matches] + mapped_matches
    ):
        workbook_fields = set(workbook_tables[workbook_table].fields)
        model_fields = set(model_tables[model_table].fields)
        table_diffs.append(
            {
                "workbook_table": workbook_table,
                "model_table": model_table,
                "match_type": "exact" if workbook_table == model_table else "alias",
                "workbook_field_count": len(workbook_fields),
                "model_field_count": len(model_fields),
                "common_count": len(workbook_fields & model_fields),
                "workbook_only": sorted(workbook_fields - model_fields),
                "model_only": sorted(model_fields - workbook_fields),
            }
        )

    table_diffs.sort(
        key=lambda item: (
            item["match_type"] != "exact",
            -len(item["workbook_only"]),
            -len(item["model_only"]),
            item["workbook_table"],
        )
    )

    return {
        "modules": {
            "python": platform.python_version(),
            "packages": collect_module_versions(["openpyxl", "pandas", "numpy"]),
        },
        "summary": {
            "workbook_table_count": len(workbook_tables),
            "model_table_count": len(model_tables),
            "exact_match_count": len(exact_matches),
            "alias_match_count": len(mapped_matches),
            "unresolved_workbook_count": len(unresolved_workbook),
            "unresolved_model_count": len(unresolved_models),
        },
        "exact_matches": exact_matches,
        "alias_matches": [
            {"workbook_table": workbook_table, "model_table": model_table}
            for workbook_table, model_table in mapped_matches
        ],
        "unresolved_workbook_tables": unresolved_workbook,
        "unresolved_model_tables": unresolved_models,
        "table_diffs": table_diffs,
    }


def collect_module_versions(module_names: list[str]) -> dict[str, dict[str, str]]:
    result: dict[str, dict[str, str]] = {}
    for module_name in module_names:
        try:
            module = importlib.import_module(module_name)
            result[module_name] = {
                "status": "ok",
                "version": str(getattr(module, "__version__", "unknown")),
            }
        except Exception as exc:  # pragma: no cover - defensive runtime branch
            result[module_name] = {"status": "error", "error": str(exc)}
    return result


def build_report(
    workbook_path: Path,
    models_path: Path,
    schema: str,
    comparison: dict[str, Any],
) -> str:
    lines = []
    summary = comparison["summary"]
    modules = comparison["modules"]
    lines.append("# SRM dictionary audit")
    lines.append("")
    lines.append(f"- Workbook: `{workbook_path}`")
    lines.append(f"- Django models: `{models_path}`")
    lines.append(f"- Schema audited: `{schema}`")
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    lines.append(
        f"- Workbook tables: {summary['workbook_table_count']} | Model tables: {summary['model_table_count']}"
    )
    lines.append(
        f"- Exact matches: {summary['exact_match_count']} | Alias matches: {summary['alias_match_count']}"
    )
    lines.append(
        f"- Workbook tables without model match: {summary['unresolved_workbook_count']}"
    )
    lines.append(
        f"- Model tables without workbook match: {summary['unresolved_model_count']}"
    )
    lines.append(f"- Python runtime: {modules['python']}")
    lines.append("")
    lines.append("## Module check")
    lines.append("")
    for module_name, details in modules["packages"].items():
        if details["status"] == "ok":
            lines.append(f"- `{module_name}`: {details['version']}")
        else:
            lines.append(f"- `{module_name}`: ERROR ({details['error']})")
    lines.append("")
    lines.append("## Alias matches")
    lines.append("")
    if comparison["alias_matches"]:
        for item in comparison["alias_matches"]:
            lines.append(
                f"- `{item['workbook_table']}` -> `{item['model_table']}`"
            )
    else:
        lines.append("- None")
    lines.append("")
    lines.append("## Workbook tables without model match")
    lines.append("")
    if comparison["unresolved_workbook_tables"]:
        for table_name in comparison["unresolved_workbook_tables"]:
            lines.append(f"- `{table_name}`")
    else:
        lines.append("- None")
    lines.append("")
    lines.append("## Model tables without workbook match")
    lines.append("")
    if comparison["unresolved_model_tables"]:
        for table_name in comparison["unresolved_model_tables"]:
            lines.append(f"- `{table_name}`")
    else:
        lines.append("- None")
    lines.append("")
    lines.append("## Field deltas")
    lines.append("")
    for diff in comparison["table_diffs"]:
        lines.append(
            f"### `{diff['workbook_table']}` -> `{diff['model_table']}` ({diff['match_type']})"
        )
        lines.append("")
        lines.append(
            f"- Workbook fields: {diff['workbook_field_count']} | Model fields: {diff['model_field_count']} | Common: {diff['common_count']}"
        )
        workbook_only_preview = ", ".join(f"`{name}`" for name in diff["workbook_only"][:12])
        model_only_preview = ", ".join(f"`{name}`" for name in diff["model_only"][:12])
        lines.append(
            f"- Workbook only ({len(diff['workbook_only'])}): {workbook_only_preview or 'None'}"
        )
        lines.append(
            f"- Model only ({len(diff['model_only'])}): {model_only_preview or 'None'}"
        )
        lines.append("")
    return "\n".join(lines)


def main() -> None:
    args = parse_args()
    workbook_tables = parse_workbook_dictionary(args.xlsx, args.schema)
    model_tables = extract_model_tables(args.models, args.schema)
    comparison = build_comparison(workbook_tables, model_tables)

    report_text = build_report(args.xlsx, args.models, args.schema, comparison)
    args.report.write_text(report_text, encoding="utf-8")
    args.json.write_text(
        json.dumps(comparison, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(report_text)
    print("")
    print(f"Markdown report written to: {args.report}")
    print(f"JSON report written to: {args.json}")


if __name__ == "__main__":
    main()
