from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


INPUT_XLSX = Path(r"C:\Users\AnasDahou\Downloads\srm_ep_only.xlsx")
OUTPUT_DIR = Path(r"C:\Users\AnasDahou\Desktop\srm_collecte\docs\ep_mobile_csvs")

CSV_COLUMNS = [
    "nom_du_champ",
    "type",
    "titre_mobile",
    "obligatoire",
    "valeur_par_defaut",
    "reference_autre_table",
    "liste_de_choix",
    "remarque_metier",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate smart EP mobile CSV specs from an SRM dictionary workbook."
    )
    parser.add_argument("--xlsx", type=Path, default=INPUT_XLSX)
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR)
    return parser.parse_args()

HEADER_NAMES = {
    "attribut": "field_name",
    "alias": "alias",
    "type": "field_type",
    "null": "nullable",
    "defaut": "default_value",
    "fk / relation": "reference",
    "description": "description",
    "liste de choix": "choice_flag",
    "valeurs possibles (lc)": "choice_values",
    "obligatoire": "required",
    "contrainte metier": "business_constraint",
    "commentaire metier": "business_comment",
}

INVALID_ALIAS_NORMALIZED = {
    "",
    "copiedeobservation",
    "anomalie_regard",
    "nom_obstacle",
    "nom_obstac",
}

GENERIC_DESCRIPTION_PREFIXES = (
    "identifiant unique",
    "fk ",
    "date/heure",
    "soft delete",
)


@dataclass
class ParsedRow:
    row_index: int
    field_name: str
    color_family: str
    color_code: str | None
    data: dict[str, str]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def normalize_key(value: Any) -> str:
    text = normalize_text(value)
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    return text.lower()


def get_fill_code(cell: openpyxl.cell.cell.Cell) -> str | None:
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return None
    color = fill.fgColor
    if color is None:
        return None
    if color.type == "rgb":
        return color.rgb
    if color.type == "indexed":
        return f"indexed:{color.indexed}"
    if color.type == "theme":
        return f"theme:{color.theme}:tint:{color.tint}"
    return str(color.type)


def rgb_tuple(fill_code: str | None) -> tuple[int, int, int] | None:
    if not fill_code or len(fill_code) != 8:
        return None
    try:
        return tuple(int(fill_code[i : i + 2], 16) for i in (2, 4, 6))
    except ValueError:
        return None


def classify_color(fill_code: str | None) -> str:
    rgb = rgb_tuple(fill_code)
    if rgb is None:
        return "none"

    r, g, b = rgb
    if min(r, g, b) >= 245 and max(r, g, b) - min(r, g, b) <= 12:
        return "white"
    if max(r, g, b) <= 80:
        return "dark"
    if g >= r + 18 and g >= b + 18:
        return "green"
    if r >= 225 and g >= 210 and b <= 215:
        return "yellow"
    if r >= 225 and g >= 170 and b >= 170:
        return "pink"
    if r >= 215 and g <= 170 and b <= 170:
        return "red"
    if b >= r + 10 and b >= g + 10:
        return "blue"
    return "other"


def is_true(value: Any) -> bool:
    raw = normalize_text(value)
    if any(symbol in raw for symbol in ("☑", "✅", "✔")):
        return True
    text = normalize_key(value)
    return text in {"true", "vrai", "oui", "1", "☑", "x"}


def is_false(value: Any) -> bool:
    raw = normalize_text(value)
    if "☐" in raw:
        return True
    text = normalize_key(value)
    return text in {"false", "faux", "non", "0", "☐"}


def parse_required(required_value: Any) -> str:
    if is_true(required_value):
        return "oui"
    return "non"


def humanize_field_name(field_name: str) -> str:
    tokens = [token for token in field_name.split("_") if token]
    if tokens and tokens[0] == "ep":
        tokens = tokens[1:]
    replacements = {
        "ref": "Reference",
        "rue": "rue",
        "num": "Numero",
        "diam": "Diametre",
        "mat": "Materiau",
        "obs": "Observation",
        "conf": "Conformite",
        "anom": "Anomalie",
        "regard": "regard",
        "pompage": "pompage",
    }
    words = []
    for token in tokens or [field_name]:
        normalized = normalize_key(token)
        words.append(replacements.get(normalized, token.capitalize()))
    title = " ".join(words)
    return re.sub(r"\s+", " ", title).strip()


def clean_alias(alias: str, field_name: str) -> str:
    if not alias:
        return humanize_field_name(field_name)

    normalized = normalize_key(alias).replace("_", "")
    if normalized in INVALID_ALIAS_NORMALIZED:
        return humanize_field_name(field_name)

    alias_clean = alias.replace("_", " ").strip()
    if alias_clean.isupper():
        alias_clean = alias_clean.title()
    return alias_clean


def normalize_choice_values(choice_values: str) -> str:
    if not choice_values:
        return ""
    choice_values = re.sub(r"\)\s*,\s*", ") | ", choice_values)
    choice_values = re.sub(r"\s+", " ", choice_values)
    return choice_values.strip()


def should_keep_description(description: str) -> bool:
    if not description:
        return False
    if re.fullmatch(r"[-+]?\d+(?:[.,]\d+)?", description):
        return False
    normalized = normalize_key(description)
    return not normalized.startswith(GENERIC_DESCRIPTION_PREFIXES)


def parse_header_map(ws: openpyxl.worksheet.worksheet.Worksheet) -> tuple[int, dict[int, str]]:
    for row_idx in range(1, min(ws.max_row, 120) + 1):
        values = [normalize_key(ws.cell(row_idx, col).value) for col in range(1, 20)]
        if "attribut" in values:
            header_map: dict[int, str] = {}
            for col_idx, value in enumerate(values, start=1):
                column_name = HEADER_NAMES.get(value)
                if column_name:
                    header_map[col_idx] = column_name
            return row_idx, header_map
    raise ValueError(f"Header row 'Attribut' not found in {ws.title}")


def dominant_row_color(ws: openpyxl.worksheet.worksheet.Worksheet, row_idx: int) -> tuple[str, str | None]:
    color_counter: Counter[str] = Counter()
    code_counter: Counter[str] = Counter()
    for col_idx in range(1, 11):
        code = get_fill_code(ws.cell(row_idx, col_idx))
        family = classify_color(code)
        if family in {"none", "white", "dark", "blue"}:
            continue
        color_counter[family] += 1
        if code:
            code_counter[code] += 1

    if not color_counter:
        return "none", None

    dominant_family = color_counter.most_common(1)[0][0]
    dominant_code = code_counter.most_common(1)[0][0] if code_counter else None
    return dominant_family, dominant_code


def is_field_candidate(values: dict[str, str]) -> bool:
    field_name = values.get("field_name", "")
    if not field_name:
        return False
    normalized = normalize_key(field_name)
    if normalized in {"schema", "table", "geometrie", "dimension", "srid", "attribut"}:
        return False
    if len(normalized) < 2:
        return False
    populated = [
        values.get("alias", ""),
        values.get("field_type", ""),
        values.get("default_value", ""),
        values.get("reference", ""),
        values.get("choice_values", ""),
        values.get("required", ""),
        values.get("business_constraint", ""),
        values.get("business_comment", ""),
    ]
    return any(populated)


def build_row(values: dict[str, str], color_family: str) -> dict[str, str]:
    field_name = values.get("field_name", "")
    alias = values.get("alias", "")
    description = values.get("description", "")
    default_value = values.get("default_value", "")
    remarks: list[str] = []

    if color_family == "green":
        remarks.append("Mode de remplissage: saisie agent mobile")
    else:
        remarks.append(
            "Mode de remplissage: pre-rempli / automatique / relation spatiale selon le code couleur source"
        )

    if default_value and normalize_key(default_value) in {"facultative", "optionnelle", "optionnel"}:
        remarks.append(f"Note: {default_value}")
        default_value = ""

    if should_keep_description(description):
        remarks.append(f"Description: {description}")

    choice_values = values.get("choice_values", "")
    choice_flag = values.get("choice_flag", "")
    list_choices = ""
    if choice_values:
        if is_true(choice_flag):
            list_choices = normalize_choice_values(choice_values)
        else:
            remarks.append(f"Valeur attendue: {choice_values}")

    business_constraint = values.get("business_constraint", "")
    business_comment = values.get("business_comment", "")
    if business_constraint:
        remarks.append(f"Contrainte: {business_constraint}")
    if business_comment:
        remarks.append(f"Commentaire: {business_comment}")

    return {
        "nom_du_champ": field_name,
        "type": values.get("field_type", ""),
        "titre_mobile": clean_alias(alias, field_name),
        "obligatoire": parse_required(values.get("required", "")),
        "valeur_par_defaut": default_value,
        "reference_autre_table": values.get("reference", ""),
        "liste_de_choix": list_choices,
        "remarque_metier": " | ".join(remarks),
    }


def parse_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> tuple[list[dict[str, str]], dict[str, Any]]:
    header_row, header_map = parse_header_map(ws)
    csv_rows: list[dict[str, str]] = []
    color_stats: Counter[str] = Counter()
    kept_fields: list[str] = []
    ignored_examples: dict[str, list[str]] = {"none": [], "white": [], "dark": []}

    for row_idx in range(header_row + 1, ws.max_row + 1):
        raw_values: dict[str, str] = {}
        for col_idx, key in header_map.items():
            raw_values[key] = normalize_text(ws.cell(row_idx, col_idx).value)

        if not is_field_candidate(raw_values):
            continue

        color_family, color_code = dominant_row_color(ws, row_idx)
        color_stats[color_family] += 1
        field_name = raw_values.get("field_name", "")

        if color_family in {"none", "white", "dark"}:
            if color_family in ignored_examples and len(ignored_examples[color_family]) < 5:
                ignored_examples[color_family].append(field_name)
            continue

        parsed = ParsedRow(
            row_index=row_idx,
            field_name=field_name,
            color_family=color_family,
            color_code=color_code,
            data=build_row(raw_values, color_family),
        )
        csv_rows.append(parsed.data)
        kept_fields.append(f"{field_name} (row {row_idx}, {color_code or 'no-code'})")

    metadata = {
        "sheet_name": ws.title,
        "table_name": normalize_text(ws.cell(4, 3).value),
        "header_row": header_row,
        "kept_count": len(csv_rows),
        "color_stats": dict(color_stats),
        "ignored_examples": {key: value for key, value in ignored_examples.items() if value},
        "kept_fields": kept_fields,
    }
    return csv_rows, metadata


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def build_markdown_summary(summary: dict[str, Any]) -> str:
    lines = [
        "# EP mobile CSV summary",
        "",
        f"- Workbook: `{summary['input_xlsx']}`",
        f"- Output folder: `{summary['output_dir']}`",
        f"- Sheets processed: {summary['sheet_count']}",
        "",
        "## By sheet",
        "",
    ]
    for sheet in summary["sheets"]:
        lines.append(f"### `{sheet['sheet_name']}`")
        lines.append("")
        lines.append(f"- Table: `{sheet['table_name']}`")
        lines.append(f"- CSV rows kept: {sheet['kept_count']}")
        lines.append(f"- Color stats seen: `{sheet['color_stats']}`")
        if sheet["ignored_examples"]:
            ignored = " | ".join(
                f"{family}: {', '.join(values)}"
                for family, values in sheet["ignored_examples"].items()
            )
            lines.append(f"- Ignored examples: {ignored}")
        if sheet["kept_fields"]:
            preview = ", ".join(sheet["kept_fields"][:8])
            lines.append(f"- Kept fields preview: {preview}")
        lines.append("")
    return "\n".join(lines)


def main() -> None:
    args = parse_args()
    output_dir = args.output_dir
    summary_json = output_dir / "_summary.json"
    summary_md = output_dir / "_summary.md"

    output_dir.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.load_workbook(args.xlsx, data_only=False)
    ep_sheets = [sheet_name for sheet_name in workbook.sheetnames if sheet_name.startswith("ep_")]

    summary = {
        "input_xlsx": str(args.xlsx),
        "output_dir": str(output_dir),
        "sheet_count": len(ep_sheets),
        "sheets": [],
    }

    for sheet_name in ep_sheets:
        ws = workbook[sheet_name]
        rows, metadata = parse_sheet(ws)
        csv_path = output_dir / f"{sheet_name}.csv"
        write_csv(csv_path, rows)
        metadata["csv_path"] = str(csv_path)
        summary["sheets"].append(metadata)

    summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    summary_md.write_text(build_markdown_summary(summary), encoding="utf-8")

    print(build_markdown_summary(summary))
    print("")
    print(f"Summary JSON written to: {summary_json}")


if __name__ == "__main__":
    main()
